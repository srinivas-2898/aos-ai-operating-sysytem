from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import os
import sys
import uuid
import json
import requests

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from main import call_deepseek

router = APIRouter()
OUTPUT_DIRECTORY = "output"
os.makedirs(OUTPUT_DIRECTORY, exist_ok=True)


class ExcelRequest(BaseModel):
    prompt: str
    sheet_type: str = "report"


def _fallback_content(prompt: str) -> dict:
    rows = [[f"Item {index}", f"Category {(index - 1) % 3 + 1}", index * 1000, index * 12.5] for index in range(1, 11)]
    return {"title": "AOS Report", "description": prompt, "primary_color": "1e3a8a", "accent_color": "3b82f6", "sheets": [
        {"sheet_name": "Overview", "description": "Project overview", "headers": ["Item", "Category", "Value", "Score"], "rows": rows, "has_chart": True, "chart_type": "bar", "chart_title": "Overview", "has_totals": True, "summary_stats": [{"label": "Records", "value": "10"}]},
        {"sheet_name": "Analysis", "description": "Detailed analysis", "headers": ["Item", "Category", "Value", "Score"], "rows": rows, "has_chart": True, "chart_type": "line", "chart_title": "Trend", "has_totals": True, "summary_stats": []},
        {"sheet_name": "Summary", "description": "Executive summary", "headers": ["Item", "Category", "Value", "Score"], "rows": rows, "has_chart": True, "chart_type": "pie", "chart_title": "Distribution", "has_totals": True, "summary_stats": []},
    ]}


def generate_excel_content(prompt: str, sheet_type: str) -> dict:
    system_prompt = """You are an expert data analyst and Excel designer. Return ONLY valid JSON containing title, description,
primary_color, accent_color, and sheets. Every sheet must include sheet_name (max 30 chars), description, headers,
rows, has_chart, chart_type (bar, line, or pie), chart_title, has_totals, and summary_stats.
Generate at least three useful sheets with at least ten meaningful rows each. Use realistic illustrative data that does not claim to be factual."""
    response = call_deepseek(system_prompt, f"Create a {sheet_type} Excel spreadsheet about: {prompt}")
    try:
        start, end = response.find("{"), response.rfind("}")
        content = json.loads(response[start:end + 1])
        if not isinstance(content.get("sheets"), list) or not content["sheets"]:
            raise ValueError("sheets missing")
        return content
    except (ValueError, json.JSONDecodeError, TypeError):
        return _fallback_content(prompt)


def _hex(value: str, fallback: str) -> str:
    value = str(value or fallback).lstrip("#")
    return value if len(value) == 6 and all(char in "0123456789abcdefABCDEF" for char in value) else fallback


def create_excel_file(content: dict, filename: str) -> str:
    filepath = os.path.join(OUTPUT_DIRECTORY, f"{filename}.xlsx")
    primary, accent = _hex(content.get("primary_color"), "1e3a8a"), _hex(content.get("accent_color"), "3b82f6")
    workbook = Workbook(); workbook.remove(workbook.active)
    border = Border(left=Side(style="thin", color="e5e7eb"), right=Side(style="thin", color="e5e7eb"), top=Side(style="thin", color="e5e7eb"), bottom=Side(style="thin", color="e5e7eb"))
    fill = lambda color: PatternFill(start_color=color, end_color=color, fill_type="solid")
    for index, sheet in enumerate(content.get("sheets", [])):
        headers = sheet.get("headers") or ["Item", "Value"]
        rows = sheet.get("rows") or []
        title = str(sheet.get("sheet_name") or f"Sheet {index + 1}")[:31]
        worksheet = workbook.create_sheet(title=title)
        last_col = get_column_letter(max(len(headers), 2))
        worksheet.merge_cells(f"A1:{last_col}1"); cell = worksheet["A1"]; cell.value = content.get("title", "Report"); cell.font = Font(name="Calibri", size=18, bold=True, color="ffffff"); cell.fill = fill(primary); cell.alignment = Alignment(horizontal="center", vertical="center"); worksheet.row_dimensions[1].height = 36
        worksheet.merge_cells(f"A2:{last_col}2"); cell = worksheet["A2"]; cell.value = sheet.get("description", content.get("description", "")); cell.font = Font(name="Calibri", size=11, color="ffffff"); cell.fill = fill(accent); cell.alignment = Alignment(indent=1); worksheet.row_dimensions[2].height = 24
        for column, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=column, value=header); cell.font = Font(name="Calibri", size=11, bold=True, color="ffffff"); cell.fill = fill(primary); cell.alignment = Alignment(horizontal="center"); cell.border = border; worksheet.column_dimensions[get_column_letter(column)].width = max(15, len(str(header)) + 6)
        for row_number, row in enumerate(rows, 5):
            for column, value in enumerate(row[:len(headers)], 1):
                cell = worksheet.cell(row=row_number, column=column, value=value); cell.font = Font(name="Calibri", size=10); cell.fill = fill("f8fafc" if row_number % 2 else "ffffff"); cell.border = border; cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left", vertical="center", indent=0 if isinstance(value, (int, float)) else 1)
                if isinstance(value, (int, float)): cell.number_format = "#,##0.00"
        end_row = len(rows) + 4
        if sheet.get("has_totals") and rows:
            total = end_row + 1; worksheet.cell(total, 1, "TOTAL").font = Font(bold=True, color="ffffff"); worksheet.cell(total, 1).fill = fill(primary)
            for column in range(2, len(headers) + 1):
                cell = worksheet.cell(total, column, f"=SUM({get_column_letter(column)}5:{get_column_letter(column)}{end_row})"); cell.font = Font(bold=True, color="ffffff"); cell.fill = fill(accent); cell.border = border; cell.number_format = "#,##0.00"
        stats = sheet.get("summary_stats") or []
        if stats:
            start = end_row + 4; worksheet.merge_cells(f"A{start}:B{start}"); worksheet.cell(start, 1, "Key Metrics").font = Font(bold=True, color="ffffff"); worksheet.cell(start, 1).fill = fill(primary)
            for offset, stat in enumerate(stats, 1):
                worksheet.cell(start + offset, 1, stat.get("label", "")).border = border; worksheet.cell(start + offset, 2, stat.get("value", "")).border = border
        if sheet.get("has_chart") and rows and len(headers) >= 2:
            chart = PieChart() if sheet.get("chart_type") == "pie" else LineChart() if sheet.get("chart_type") == "line" else BarChart()
            chart.title = sheet.get("chart_title", "Chart"); chart.style = 10; chart.width = 20; chart.height = 14
            chart.add_data(Reference(worksheet, min_col=2, max_col=min(4, len(headers)), min_row=4, max_row=end_row), titles_from_data=True)
            chart.set_categories(Reference(worksheet, min_col=1, min_row=5, max_row=end_row)); worksheet.add_chart(chart, f"A{end_row + 8}")
        worksheet.freeze_panes = "A5"
    workbook.save(filepath)
    return filepath


@router.post("/api/generate-excel")
async def generate_excel(request: ExcelRequest):
    try:
        content = generate_excel_content(request.prompt, request.sheet_type)
        filepath = create_excel_file(content, f"aos_excel_{uuid.uuid4().hex[:8]}")
        title = "".join(char for char in str(content.get("title", "spreadsheet")) if char.isalnum() or char in (" ", "-", "_")) or "spreadsheet"
        return FileResponse(filepath, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"{title}.xlsx")
    except HTTPException:
        raise
    except Exception as error:
        raise HTTPException(status_code=500, detail=str(error)) from error
